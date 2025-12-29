import json
import unicodedata
import openpyxl

EXCEL_FILE = "Classification.xlsx"
SHEET_NAME = "Base"
OUTPUT_FILE = "base-data.json"


def norm(s: str) -> str:
    if s is None:
        return ""
    s = str(s).strip().lower()
    s = unicodedata.normalize("NFD", s)
    return "".join(ch for ch in s if unicodedata.category(ch) != "Mn")


def main():
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Sheet '{SHEET_NAME}' not found in workbook")

    sh = wb[SHEET_NAME]

    values = [list(row) for row in sh.iter_rows(values_only=True)]

    if len(values) < 2:
        raise ValueError("Base sheet must have at least 2 header rows")

    header_row_2 = values[1]
    apercu_index = None
    for i, val in enumerate(header_row_2):
        if norm(val) == "apercu":
            apercu_index = i
            break

    if apercu_index is None:
        print(
            "Warning: no 'Apercu' column found in second header row "
            "(expected something like 'Apercu')"
        )
    else:
        print(f"'Apercu' column detected at index {apercu_index} (0-based).")

    data = {"values": values}

    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False)

    print(f"Written {len(values)} rows to {OUTPUT_FILE}")


if __name__ == "__main__":
    main()

